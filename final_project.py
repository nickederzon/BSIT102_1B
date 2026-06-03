import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op

window = tk.Tk()
window.title("First Aid Supply Inventory System")
window.configure(bg="lightcyan")


def display():
    workbook = op.load_workbook("QUEBRADO_database.xlsx")
    sheet = workbook.active

    for content in table.get_children():
        table.delete(content)

    for rows in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=rows)


def validation():
    supply_name = supply_entry.get()
    category = category_entry.get()
    quantity = qty_entry.get()
    unit_price = price_entry.get()

    if not supply_name or not category or not quantity or not unit_price:
        messagebox.showerror("Error", "Please fill in all fields.")
        return False

    if not quantity.isdigit() or not unit_price.isdigit():
        messagebox.showerror("Error", "Quantity and Unit Price must be numbers only.")
        return False

    return True


def create():
    if not validation():
        return

    supply_name = supply_entry.get()
    category = category_entry.get()
    quantity = int(qty_entry.get())
    unit_price = int(price_entry.get())

    total_value = quantity * unit_price

    workbook = op.load_workbook("QUEBRADO_database.xlsx")
    sheet = workbook.active

    supply_id = sheet.max_row

    sheet.append([supply_id, supply_name, category, quantity, unit_price, total_value])

    workbook.save("QUEBRADO_database.xlsx")

    messagebox.showinfo("Success", "Supply added successfully!")

    clear()
    display()


def auto_populate(event):
    selected = table.focus()
    values = table.item(selected, "values")

    if values:
        supply_entry.delete(0, tk.END)
        category_entry.delete(0, tk.END)
        qty_entry.delete(0, tk.END)
        price_entry.delete(0, tk.END)

        supply_entry.insert(0, values[1])
        category_entry.insert(0, values[2])
        qty_entry.insert(0, values[3])
        price_entry.insert(0, values[4])


def update():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a supply first.")
        return

    if not validation():
        return

    supply_name = supply_entry.get()
    category = category_entry.get()
    quantity = int(qty_entry.get())
    unit_price = int(price_entry.get())

    total_value = quantity * unit_price

    values = table.item(selected, "values")
    supply_id = values[0]

    workbook = op.load_workbook("QUEBRADO_database.xlsx")
    sheet = workbook.active

    for rows in sheet.iter_rows(min_row=2):
        if str(supply_id) == str(rows[0].value):
            rows[1].value = supply_name
            rows[2].value = category
            rows[3].value = quantity
            rows[4].value = unit_price
            rows[5].value = total_value

    workbook.save("QUEBRADO_database.xlsx")

    messagebox.showinfo("Success", "Supply updated successfully!")

    clear()
    display()


def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a supply first.")
        return

    confirm = messagebox.askyesno("Confirm","Are you sure you want to delete this supply?" )

    if not confirm:
        return

    values = table.item(selected, "values")
    supply_id = values[0]

    workbook = op.load_workbook("QUEBRADO_database.xlsx")
    sheet = workbook.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(supply_id) == str(row[0].value):
            sheet.delete_rows(i)
            break

    workbook.save("QUEBRADO_database.xlsx")

    messagebox.showinfo("Success", "Supply deleted successfully!")

    clear()
    display()


def clear():
    supply_entry.delete(0, tk.END)
    category_entry.delete(0, tk.END)
    qty_entry.delete(0, tk.END)
    price_entry.delete(0, tk.END)


# Title
title = tk.Label(window,text="First Aid Supply Inventory System",font=("Times New Roman", 16, "bold"),bg="lightcyan")
title.grid(row=0, column=0, columnspan=6, pady=10)


# Frame
genframe = tk.Frame(window,bg="lightcyan",bd=2,relief="groove")
genframe.grid(row=1, column=0, columnspan=7, padx=10, pady=10)


# Supply Name
supply_entry = tk.Entry(genframe, font=("Poppins", 12))
supply_entry.grid(row=2, column=1, columnspan=2, padx=10, pady=(10, 0))

supply_label = tk.Label(genframe,text="Supply Name", font=("Poppins", 10, "italic"),bg="lightcyan")
supply_label.grid(row=3, column=1, columnspan=2)


# Category
category_entry = tk.Entry(genframe, font=("Poppins", 12))
category_entry.grid(row=2, column=3, columnspan=2, padx=10, pady=(10, 0))

category_label = tk.Label(genframe,text="Category",font=("Poppins", 10, "italic"),bg="lightcyan")
category_label.grid(row=3, column=3, columnspan=2)


# Quantity
qty_entry = tk.Entry(genframe, font=("Poppins", 12))
qty_entry.grid(row=4, column=1, columnspan=2, padx=10, pady=(10, 0))

qty_label = tk.Label(genframe,text="Quantity",font=("Poppins", 10, "italic"),bg="lightcyan")
qty_label.grid(row=5, column=1, columnspan=2)


# Unit Price
price_entry = tk.Entry(genframe, font=("Poppins", 12))
price_entry.grid(row=4, column=3, columnspan=2, padx=10, pady=(10, 0))

price_label = tk.Label(genframe,text="Unit Price",font=("Poppins", 10, "italic"),bg="lightcyan")
price_label.grid(row=5, column=3, columnspan=2)


# Buttons
submit_btn = tk.Button(window,text="Add Supply",font=("Poppins", 12, "bold"),bg="lightgreen",command=create)
submit_btn.grid(row=6, column=1, pady=(10, 20))

update_btn=tk.Button(window,text="Update",font=("Poppins",12,"bold"),bg="lightyellow",command=update)update_btn.grid(row=6, column=2)

delete_btn=tk.Button(window,text="Delete",font=("Poppins",12,"bold"),bg="lightcoral",command=delete)delete_btn.grid(row=6, column=3)


# Table
table = ttk.Treeview(window,columns=("Supply ID","Supply Name","Category","Quantity","Unit Price","Total Value"),show="headings")

for heading in ("Supply ID","Supply Name","Category","Quantity","Unit Price","Total Value"):
    table.heading(heading, text=heading)

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)

table.bind("<<TreeviewSelect>>", auto_populate)

display()

window.mainloop()